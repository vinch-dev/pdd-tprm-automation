# -*- coding: utf-8 -*-
"""
Process: SAP Automated Procurement Report Extraction
Description: Extracts historical PO cost data based on project configurations.
Version: 1.0 (Production)
"""

import win32com.client
import time
import re
from pathlib import Path
from openpyxl import load_workbook
from datetime import date

# ==============================================================================
# 1. CONFIGURATION & DIRECTORY SETUP
# ==============================================================================
# Standardized paths for Enterprise Environment deployment
base_dir = Path(__file__).parent.resolve()
config_excel_path = base_dir / "Processing_Configuration_Matrix.xlsx"
export_output_folder = base_dir / "SAP_Extracted_Reports"

config_sheet_name = "Selection_Criteria"

if not config_excel_path.exists():
    print(f"Error: Configuration file missing at: {config_excel_path}")
    raise SystemExit

# Ensure output directory exists
export_output_folder.mkdir(parents=True, exist_ok=True)

# ==============================================================================
# 2. DATA INGESTION (Configuration Matrix)
# ==============================================================================
print(f"Loading selection criteria from: {config_excel_path.name}")

try:
    wb = load_workbook(config_excel_path, data_only=True)
    ws = wb[config_sheet_name]
except KeyError:
    print(f"Error: Required sheet '{config_sheet_name}' not found.")
    raise SystemExit

def get_clean_val(cell_ref):
    """Helper to extract and sanitize Excel cell values."""
    val = ws[cell_ref].value
    return str(val).strip() if val is not None else ""

# --- Mapping Configuration Fields ---
# Purchasing Groups (Max 8)
PUR_GRP_MAIN = get_clean_val("B4")
PUR_GRP_LIST = [get_clean_val(c) for c in ["C4", "D4", "E4", "F4", "G4", "H4", "I4"]]

# Plant Codes
PLANTS = [get_clean_val(c) for c in ["B5", "C5", "D5"]]

# Project Identifiers
PROJ_MAIN = get_clean_val("B6")
PROJ_EXT = [get_clean_val(c) for c in ["C6", "D6", "E6", "F6"]]

# Vendor IDs
VENDOR_MAIN = get_clean_val("B8")
VENDOR_LIST = [get_clean_val(c) for c in ["C8", "D8", "E8", "F8", "G8", "H8", "I8"]]

# Document Numbers (PO)
DOC_MAIN = get_clean_val("B7")
DOC_LIST = [get_clean_val(c) for c in ["C7", "D7", "E7", "F7", "G7"]]

# Filters
FILTER_TEXT = get_clean_val("B11")

# ==============================================================================
# 3. DATE CALCULATION LOGIC (Last Calendar Week)
# ==============================================================================
today = date.today()
iso_year, current_wk, _ = today.isocalendar()

# Target the previous completed calendar week
target_wk = current_wk - 1
start_date = date.fromisocalendar(iso_year, target_wk, 1).strftime("%d.%m.%Y")
end_date   = date.fromisocalendar(iso_year, target_wk, 7).strftime("%d.%m.%Y")

# Define Output Filename
clean_proj_name = re.sub(r'[\\/*?:"<>|]', "", PROJ_MAIN)
output_filename = f"Procurement_Report_{clean_proj_name}_Wk{target_wk}_{iso_year}.xlsx"
full_save_path = export_output_folder / output_filename

# ==============================================================================
# 4. SAP GUI CONNECTIVITY
# ==============================================================================
try:
    sap_gui = win32com.client.GetObject("SAPGUI")
    application = sap_gui.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)
except Exception as e:
    print("Connectivity Error: Ensure SAP GUI is active and user is logged in.")
    raise SystemExit

# ==============================================================================
# 5. SAP AUTOMATION LOGIC
# ==============================================================================
try:
    # Initialize Transaction
    session.findById("wnd[0]/tbar[0]/okcd").text = "INTERNAL_REPORT_NAME"
    session.findById("wnd[0]").sendVKey(0)

    # --- Input Selection Criteria ---
    session.findById("wnd[0]/usr/ctxtSO_EBELN-LOW").text = DOC_MAIN # PO Number
    session.findById("wnd[0]/usr/ctxtSO_WERKS-LOW").text = PLANTS[0] # Primary Plant
    
    # Handle Multi-Selection for Plants
    if any(PLANTS[1:]):
        session.findById("wnd[0]/usr/btn%_SO_WERKS_%_APP_%-VALU_PUSH").press()
        for idx, val in enumerate(PLANTS[1:], start=1):
            if val:
                session.findById(f"wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,{idx}]").text = val
        session.findById("wnd[1]").sendVKey(8)

    # Handle Multi-Selection for Document Numbers
    if any(DOC_LIST):
        session.findById("wnd[0]/usr/btn%_SO_EBELN_%_APP_%-VALU_PUSH").press()
        for idx, val in enumerate(DOC_LIST):
            if val:
                session.findById(f"wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,{idx+1}]").text = val
        session.findById("wnd[1]").sendVKey(8)

    # Fill Purchasing Groups
    session.findById("wnd[0]/usr/ctxtSO_EKGRP-LOW").text = PUR_GRP_MAIN
    if any(PUR_GRP_LIST):
        session.findById("wnd[0]/usr/btn%_SO_EKGRP_%_APP_%-VALU_PUSH").press()
        for idx, val in enumerate(PUR_GRP_LIST):
            if val:
                session.findById(f"wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,{idx+1}]").text = val
        session.findById("wnd[1]").sendVKey(8)

    # Set Reporting Dates
    session.findById("wnd[0]/usr/ctxtSO_EBDAT-LOW").text = start_date
    session.findById("wnd[0]/usr/ctxtSO_EBDAT-HIGH").text = end_date

    # --- Execution & Layout Handling ---
    print("Executing SAP Data Retrieval...")
    time.sleep(2)
    session.findById("wnd[0]").sendVKey(8) # Execute F8
    
    # --- Export to Spreadsheet ---
    print(f"Exporting results to: {output_filename}")
    session.findById("wnd[0]").sendVKey(46) # Export List
    session.findById("wnd[0]").sendVKey(43) # Choose Spreadsheet
    session.findById("wnd[1]").sendVKey(0)  # Enter/Confirm

    # SAP Scripting requires string conversion for Path objects
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = str(export_output_folder)
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = output_filename
    
    session.findById("wnd[1]").sendVKey(11) # Save/Replace

    # Return to Main Menu
    for _ in range(2):
        session.findById("wnd[0]").sendVKey(12)

    print("Automation Cycle Completed Successfully.")

except Exception as e:
    print(f"Business Exception / SAP Scripting Error: {e}")
